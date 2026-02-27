#!/usr/bin/env python3
import argparse
import csv
import io
import json
import openpyxl
import os
import re
import hmac
import base64
import time
import hashlib
import secrets
from datetime import datetime, timedelta
from collections import defaultdict
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from urllib import request

# Optional document extraction libraries
try:
    import pypdf
    _HAVE_PYPDF = True
except ImportError:
    _HAVE_PYPDF = False

try:
    import docx as _docx
    _HAVE_DOCX = True
except ImportError:
    _HAVE_DOCX = False

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data" / "processed"
EDITS_FILE = DATA / "roadmap_edits.json"
CLUSTER_EDITS_FILE = DATA / "cluster_edits.json"
USERS_FILE = DATA / "users.json"
ACTIVITY_FILE = DATA / "activity_log.json"
INVITES_FILE = DATA / "invites.json"
AUTH_DISABLED = os.environ.get("AUTH_DISABLED", "false").strip().lower() == "true"
SESSION_SECRET = os.environ.get("SESSION_SECRET", "dev-insecure-change-me")
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "")
ALLOWED_DOMAIN = os.environ.get("ALLOWED_SSO_DOMAIN", "clear.in").strip().lower()
SESSION_COOKIE = "ccc_session"
DELETE_ADMIN_EMAIL = os.environ.get("DELETE_ADMIN_EMAIL", "aditya.jaiswal@clear.in").strip().lower()
ADMIN_BOOTSTRAP_EMAIL = os.environ.get("ADMIN_BOOTSTRAP_EMAIL", "aditya.jaiswal@clear.in").strip().lower()
ROLE_ADMIN = "admin"
ROLE_EDIT = "edit"
ROLE_VIEW = "view"

DONE_SET = {"done", "closed", "completed"}
PROGRESS_SET = {"in progress", "problem solving", "deferred", "at risk", "triage"}
TODO_SET = {"to do", "not started", "", "backlog", "won't do"}
QUARTERS = ["Q1", "Q2", "Q3", "Q4", "Unassigned"]


def b64url_encode(data_bytes):
    return base64.urlsafe_b64encode(data_bytes).decode("ascii").rstrip("=")


def b64url_decode(data_str):
    pad = "=" * ((4 - len(data_str) % 4) % 4)
    return base64.urlsafe_b64decode((data_str + pad).encode("ascii"))


def sign_payload(payload_b64):
    mac = hmac.new(SESSION_SECRET.encode("utf-8"), payload_b64.encode("ascii"), hashlib.sha256).digest()
    return b64url_encode(mac)


def issue_session(email):
    now = int(time.time())
    body = {"email": email, "exp": now + (8 * 60 * 60)}
    payload_b64 = b64url_encode(json.dumps(body, ensure_ascii=True, separators=(",", ":")).encode("utf-8"))
    sig = sign_payload(payload_b64)
    return f"{payload_b64}.{sig}"


def verify_session(token):
    if not token or "." not in token:
        return None
    payload_b64, sig = token.split(".", 1)
    expected = sign_payload(payload_b64)
    if not hmac.compare_digest(sig, expected):
        return None
    try:
        body = json.loads(b64url_decode(payload_b64).decode("utf-8"))
    except Exception:
        return None
    exp = int(body.get("exp", 0) or 0)
    if exp < int(time.time()):
        return None
    email = normalize_email(body.get("email"))
    if not email:
        return None
    return body


def parse_cookies(cookie_header):
    out = {}
    if not cookie_header:
        return out
    parts = cookie_header.split(";")
    for p in parts:
        if "=" not in p:
            continue
        k, v = p.split("=", 1)
        out[k.strip()] = v.strip()
    return out


def load_users():
    if not USERS_FILE.exists():
        return {"users": {}}
    try:
        data = json.loads(USERS_FILE.read_text(encoding="utf-8"))
        data.setdefault("users", {})
        return data
    except Exception:
        return {"users": {}}


def save_users(data):
    USERS_FILE.write_text(json.dumps(data, ensure_ascii=True, indent=2), encoding="utf-8")


def load_activity():
    if not ACTIVITY_FILE.exists():
        return {"events": []}
    try:
        data = json.loads(ACTIVITY_FILE.read_text(encoding="utf-8"))
        data.setdefault("events", [])
        return data
    except Exception:
        return {"events": []}


def save_activity(data):
    save_count = 500
    data["events"] = (data.get("events", []) or [])[-save_count:]
    ACTIVITY_FILE.write_text(json.dumps(data, ensure_ascii=True, indent=2), encoding="utf-8")


def load_invites():
    if not INVITES_FILE.exists():
        return {"invites": {}}
    try:
        data = json.loads(INVITES_FILE.read_text(encoding="utf-8"))
        data.setdefault("invites", {})
        return data
    except Exception:
        return {"invites": {}}


def save_invites(data):
    INVITES_FILE.write_text(json.dumps(data, ensure_ascii=True, indent=2), encoding="utf-8")


def normalize_email(email):
    return (email or "").strip().lower()


def normalize_role(role):
    r = (role or "").strip().lower()
    if r in {ROLE_ADMIN, ROLE_EDIT, ROLE_VIEW}:
        return r
    return ROLE_VIEW


def role_can_edit(role):
    return normalize_role(role) in {ROLE_ADMIN, ROLE_EDIT}


def role_can_delete(role):
    return normalize_role(role) == ROLE_ADMIN


def role_can_manage_users(role):
    return normalize_role(role) == ROLE_ADMIN


def email_is_valid(email):
    e = normalize_email(email)
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", e))


def hash_password(password, salt=None):
    s = salt or secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), s.encode("utf-8"), 120000)
    return {"salt": s, "hash": b64url_encode(dk)}


def verify_password(password, salt, expected_hash):
    got = hash_password(password, salt=salt)["hash"]
    return hmac.compare_digest(got, expected_hash or "")


def get_user_record(email):
    e = normalize_email(email)
    if not e:
        return None
    users = load_users()
    user = users.get("users", {}).get(e)
    if not user:
        return None
    role = normalize_role(user.get("role", ROLE_VIEW))
    user["role"] = role
    return user


def ensure_bootstrap_admin(email):
    e = normalize_email(email)
    if not e:
        return
    users = load_users()
    users.setdefault("users", {})
    u = users["users"].get(e)
    changed = False
    if not u:
        u = {"email": e, "created_at": int(time.time())}
        users["users"][e] = u
        changed = True
    if e == ADMIN_BOOTSTRAP_EMAIL and normalize_role(u.get("role")) != ROLE_ADMIN:
        u["role"] = ROLE_ADMIN
        changed = True
    elif "role" not in u:
        u["role"] = ROLE_VIEW
        changed = True
    if changed:
        save_users(users)


def apply_invite_to_user(email, invite_token):
    token = (invite_token or "").strip()
    if not token:
        return None
    invites = load_invites()
    inv = invites.get("invites", {}).get(token)
    if not inv:
        return None
    if inv.get("status") != "pending":
        return None
    invited_email = normalize_email(inv.get("email"))
    e = normalize_email(email)
    if invited_email and invited_email != e:
        return None
    role = normalize_role(inv.get("role", ROLE_VIEW))
    users = load_users()
    users.setdefault("users", {})
    if e in users["users"]:
        users["users"][e]["role"] = role
        save_users(users)
    inv["status"] = "accepted"
    inv["accepted_at"] = int(time.time())
    inv["accepted_by"] = e
    invites["invites"][token] = inv
    save_invites(invites)
    return inv


def verify_google_id_token(id_token):
    if not id_token:
        return {"ok": False, "reason": "Missing id_token"}
    q = f"https://oauth2.googleapis.com/tokeninfo?id_token={id_token}"
    req = request.Request(q, headers={"User-Agent": "ccc-dashboard-auth/1.0"})
    try:
        with request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        return {"ok": False, "reason": f"Token verification failed: {e}"}

    aud = (data.get("aud") or "").strip()
    email = (data.get("email") or "").strip().lower()
    email_verified = str(data.get("email_verified", "")).lower() == "true"
    hd = (data.get("hd") or "").strip().lower()

    if GOOGLE_CLIENT_ID and aud != GOOGLE_CLIENT_ID:
        return {"ok": False, "reason": "Invalid client id"}
    if not email_verified:
        return {"ok": False, "reason": "Email not verified by Google"}
    if not email.endswith(f"@{ALLOWED_DOMAIN}"):
        return {"ok": False, "reason": f"Only {ALLOWED_DOMAIN} accounts are allowed"}
    if hd and hd != ALLOWED_DOMAIN:
        return {"ok": False, "reason": f"Hosted domain must be {ALLOWED_DOMAIN}"}
    return {"ok": True, "email": email}


def extract_text_from_document(data_bytes, filename, mime_type=""):
    """Extract plain text from uploaded document bytes.
    Supports PDF, DOCX, XLSX, XLS, CSV, TXT, and plain text files."""
    name_lower = (filename or "").lower()
    mime_lower = (mime_type or "").lower()

    # ── PDF ──────────────────────────────────────────────────────────────────
    if name_lower.endswith(".pdf") or "pdf" in mime_lower:
        if not _HAVE_PYPDF:
            return "", "pypdf library not installed"
        try:
            reader = pypdf.PdfReader(io.BytesIO(data_bytes))
            pages = []
            for page in reader.pages:
                t = page.extract_text() or ""
                if t.strip():
                    pages.append(t.strip())
            text = "\n\n".join(pages)
            return text[:40000], None  # cap at 40k chars
        except Exception as ex:
            return "", f"PDF read error: {ex}"

    # ── DOCX ─────────────────────────────────────────────────────────────────
    if name_lower.endswith(".docx") or "wordprocessingml" in mime_lower or "vnd.openxmlformats-officedocument.wordprocessingml" in mime_lower:
        if not _HAVE_DOCX:
            return "", "python-docx library not installed"
        try:
            doc = _docx.Document(io.BytesIO(data_bytes))
            paras = [p.text for p in doc.paragraphs if p.text.strip()]
            text = "\n".join(paras)
            return text[:40000], None
        except Exception as ex:
            return "", f"DOCX read error: {ex}"

    # ── XLSX / XLS ────────────────────────────────────────────────────────────
    if name_lower.endswith((".xlsx", ".xls")) or "spreadsheetml" in mime_lower or "ms-excel" in mime_lower:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
            lines = []
            for ws in wb.worksheets:
                lines.append(f"[Sheet: {ws.title}]")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in cells):
                        lines.append("\t".join(cells))
            text = "\n".join(lines)
            return text[:40000], None
        except Exception as ex:
            return "", f"XLSX read error: {ex}"

    # ── CSV ───────────────────────────────────────────────────────────────────
    if name_lower.endswith(".csv") or "text/csv" in mime_lower:
        try:
            text = data_bytes.decode("utf-8", errors="replace")
            return text[:40000], None
        except Exception as ex:
            return "", f"CSV read error: {ex}"

    # ── Plain text / TXT / markdown ───────────────────────────────────────────
    if name_lower.endswith((".txt", ".md", ".rst")) or mime_lower.startswith("text/"):
        try:
            text = data_bytes.decode("utf-8", errors="replace")
            return text[:40000], None
        except Exception as ex:
            return "", f"Text read error: {ex}"

    # ── PPT / PPTX — return filename only (no parser available) ───────────────
    return "", f"Unsupported file type: {name_lower.split('.')[-1] if '.' in name_lower else 'unknown'}"


def load_csv(path):
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as fh:
        return list(csv.DictReader(fh))


def load_edits():
    if not EDITS_FILE.exists():
        return {"rows": {}, "added_rows": {}, "deleted_rows": {}}
    try:
        loaded = json.loads(EDITS_FILE.read_text(encoding="utf-8"))
        loaded.setdefault("rows", {})
        loaded.setdefault("added_rows", {})
        loaded.setdefault("deleted_rows", {})
        return loaded
    except Exception:
        return {"rows": {}, "added_rows": {}, "deleted_rows": {}}


def save_edits(data):
    EDITS_FILE.write_text(json.dumps(data, ensure_ascii=True, indent=2), encoding="utf-8")


def load_cluster_edits():
    if not CLUSTER_EDITS_FILE.exists():
        return {"manual_rows": {}}
    try:
        data = json.loads(CLUSTER_EDITS_FILE.read_text(encoding="utf-8"))
        data.setdefault("manual_rows", {})
        return data
    except Exception:
        return {"manual_rows": {}}


def save_cluster_edits(data):
    CLUSTER_EDITS_FILE.write_text(json.dumps(data, ensure_ascii=True, indent=2), encoding="utf-8")


def apply_row_edits(items, edits):
    rows_map = edits.get("rows", {})
    added_rows = edits.get("added_rows", {})
    deleted_rows = edits.get("deleted_rows", {})
    out = []
    for item in items:
        row_id = item.get("roadmap_item_id", "")
        if row_id in deleted_rows:
            continue
        patch = rows_map.get(row_id, {})
        merged = dict(item)
        for k, v in patch.items():
            merged[k] = v
        out.append(normalize_runtime_item(merged))

    for row_id, raw_row in (added_rows or {}).items():
        if row_id in deleted_rows:
            continue
        merged = {"roadmap_item_id": row_id}
        if isinstance(raw_row, dict):
            merged.update(raw_row)
        patch = rows_map.get(row_id, {})
        if isinstance(patch, dict):
            merged.update(patch)
        out.append(normalize_runtime_item(merged))

    return out


def status_to_progress(status):
    s = (status or "").strip().lower()
    if s in DONE_SET:
        return 100.0
    if s in PROGRESS_SET:
        return 50.0
    if s in TODO_SET:
        return 0.0
    return 25.0


def normalize_quarter(value, end_date):
    s = (value or "").upper()
    m = re.search(r"Q[1-4]", s)
    if m:
        return m.group(0)

    d = (end_date or "").upper()
    month_map = {
        "JAN": "Q1", "FEB": "Q1", "MAR": "Q1",
        "APR": "Q2", "MAY": "Q2", "JUN": "Q2",
        "JUL": "Q3", "AUG": "Q3", "SEP": "Q3",
        "OCT": "Q4", "NOV": "Q4", "DEC": "Q4",
    }
    for mon, q in month_map.items():
        if mon in d:
            return q
    return "Unassigned"


def normalize_item(row):
    status_raw = (row.get("status") or "").strip()
    status = status_raw.lower()
    try:
        progress = float(row.get("percent_complete") or 0)
    except ValueError:
        progress = 0
    if progress <= 0 and status:
        progress = status_to_progress(status)

    quarter = normalize_quarter(row.get("target_quarter", ""), row.get("end_date", ""))

    return {
        "roadmap_item_id": row.get("roadmap_item_id", ""),
        "product": (row.get("product") or "Unknown").strip(),
        "hierarchy_level": (row.get("hierarchy_level") or "").strip(),
        "parent_id": (row.get("parent_id") or "").strip(),
        "stack_rank": (row.get("stack_rank") or "").strip(),
        "theme": (row.get("theme") or "Uncategorized").strip() or "Uncategorized",
        "title": (row.get("title") or "Untitled").strip(),
        "description": (row.get("description") or "").strip(),
        "aop_goal": (row.get("aop_goal") or "").strip(),
        "priority": (row.get("priority") or "").strip(),
        "status_raw": status_raw,
        "status": status,
        "progress": progress,
        "owner": (row.get("owner") or "Unassigned").strip() or "Unassigned",
        "quarter": quarter,
        "start_date": (row.get("start_date") or "").strip(),
        "end_date": (row.get("end_date") or "").strip(),
        "prod_date_projected": (row.get("prod_date_projected") or "").strip(),
        "old_prod_date_projected": (row.get("old_prod_date_projected") or "").strip(),
        "customer_facing_date": (row.get("customer_facing_date") or "").strip(),
        "source_system": (row.get("source_system") or "").strip(),
        "source_key": (row.get("source_key") or "").strip(),
        "source_file": (row.get("source_file") or "").strip(),
        "prd_generated": (row.get("prd_generated") or "").strip(),
        "prd_description": (row.get("prd_description") or "").strip(),
        "prd_video_link": (row.get("prd_video_link") or "").strip(),
        "prd_customer_doc_link": (row.get("prd_customer_doc_link") or "").strip(),
        "prd_use_doc_inference": (row.get("prd_use_doc_inference") or "").strip(),
        "prd_generated_at": (row.get("prd_generated_at") or "").strip(),
        "prd_content": (row.get("prd_content") or "").strip(),
    }


def normalize_runtime_item(row):
    status_raw = (row.get("status_raw") or row.get("status") or "").strip()
    status = status_raw.lower()
    try:
        progress = float(row.get("progress") or row.get("percent_complete") or 0)
    except ValueError:
        progress = 0.0
    if progress <= 0 and status:
        progress = status_to_progress(status)

    quarter = (row.get("quarter") or "").strip()
    if not quarter:
        quarter = normalize_quarter(row.get("target_quarter", ""), row.get("end_date", ""))
    quarter = quarter or "Unassigned"

    return {
        "roadmap_item_id": (row.get("roadmap_item_id") or "").strip(),
        "product": (row.get("product") or "Unknown").strip(),
        "hierarchy_level": (row.get("hierarchy_level") or "Initiative").strip() or "Initiative",
        "parent_id": (row.get("parent_id") or "").strip(),
        "stack_rank": (row.get("stack_rank") or "").strip(),
        "theme": (row.get("theme") or "Uncategorized").strip() or "Uncategorized",
        "title": (row.get("title") or "Untitled").strip(),
        "description": (row.get("description") or "").strip(),
        "aop_goal": (row.get("aop_goal") or "").strip(),
        "priority": (row.get("priority") or "").strip(),
        "status_raw": status_raw,
        "status": status,
        "progress": progress,
        "owner": (row.get("owner") or "Unassigned").strip() or "Unassigned",
        "quarter": quarter,
        "start_date": (row.get("start_date") or "").strip(),
        "end_date": (row.get("end_date") or "").strip(),
        "prod_date_projected": (row.get("prod_date_projected") or "").strip(),
        "old_prod_date_projected": (row.get("old_prod_date_projected") or "").strip(),
        "customer_facing_date": (row.get("customer_facing_date") or "").strip(),
        "source_system": (row.get("source_system") or "Manual").strip(),
        "source_key": (row.get("source_key") or "").strip(),
        "source_file": (row.get("source_file") or "UI").strip(),
        "prd_generated": (row.get("prd_generated") or "").strip(),
        "prd_description": (row.get("prd_description") or "").strip(),
        "prd_video_link": (row.get("prd_video_link") or "").strip(),
        "prd_customer_doc_link": (row.get("prd_customer_doc_link") or "").strip(),
        "prd_use_doc_inference": (row.get("prd_use_doc_inference") or "").strip(),
        "prd_generated_at": (row.get("prd_generated_at") or "").strip(),
        "prd_content": (row.get("prd_content") or "").strip(),
    }


def avg_progress(rows):
    if not rows:
        return 0.0
    return sum(float(r.get("progress", 0) or 0) for r in rows) / len(rows)


def normalize_query_value(params, key, default="ALL"):
    return params.get(key, [default])[0]


def filter_items(items, product="ALL", quarter="ALL", status="ALL"):
    out = []
    for r in items:
        if product != "ALL" and r["product"] != product:
            continue
        if quarter != "ALL" and r["quarter"] != quarter:
            continue
        if status != "ALL" and r["status_raw"] != status:
            continue
        out.append(r)
    return out


def roadmap_summary(items, product="ALL"):
    scoped = filter_items(items, product=product)
    total = len(scoped)
    done = sum(1 for r in scoped if r["status"] in DONE_SET)
    in_progress = sum(1 for r in scoped if r["status"] in PROGRESS_SET)
    todo = sum(1 for r in scoped if r["status"] in TODO_SET)
    return {
        "scope": product,
        "annual_progress": round(avg_progress(scoped), 2),
        "total_items": total,
        "done_items": done,
        "in_progress_items": in_progress,
        "todo_items": todo,
    }


def roadmap_quarterly(items, product="ALL"):
    scoped = filter_items(items, product=product)
    by_scope = defaultdict(list)
    for r in scoped:
        by_scope[r["quarter"]].append(r)

    return [
        {
            "quarter": q,
            "items": len(by_scope[q]),
            "progress": round(avg_progress(by_scope[q]), 2) if by_scope[q] else None,
        }
        for q in QUARTERS
    ]


def product_rollups(items):
    products = sorted({x["product"] for x in items})
    out = []
    for p in products:
        rows = filter_items(items, product=p)
        out.append({
            "product": p,
            "items": len(rows),
            "progress": round(avg_progress(rows), 2),
        })
    return out


def to_int(value, default=0):
    try:
        return int(float(value or 0))
    except Exception:
        return default


def enrich_clusters_with_customers(cluster_rows, assignment_rows, request_rows):
    if not cluster_rows:
        return []

    req_by_id = {}
    for r in request_rows or []:
        # salesforce_requests_unified.csv uses "Request ID" (Title Case with space)
        req_id = (r.get("Request ID") or r.get("request_id") or "").strip()
        if req_id:
            req_by_id[req_id] = r

    customers_by_cluster = defaultdict(dict)
    for a in assignment_rows or []:
        cluster_id = (a.get("cluster_id") or "").strip()
        req_id = (a.get("request_id") or "").strip()
        if not cluster_id or not req_id:
            continue
        req = req_by_id.get(req_id, {})
        # salesforce_requests_unified.csv uses "Account Name" and "Ticket Count" (Title Case with spaces)
        name = (req.get("Account Name") or req.get("account_name") or "").strip() or "Unknown Customer"
        ticket_count = to_int(req.get("Ticket Count") or req.get("ticket_count"), default=1)
        bucket = customers_by_cluster[cluster_id].setdefault(name, {"name": name, "tickets": 0, "requests": 0})
        bucket["tickets"] += max(ticket_count, 1)
        bucket["requests"] += 1

    out = []
    for row in cluster_rows:
        cluster_id = (row.get("cluster_id") or "").strip()
        customers = list(customers_by_cluster.get(cluster_id, {}).values())
        customers.sort(key=lambda x: (-x["tickets"], -x["requests"], x["name"]))
        names = [c["name"] for c in customers]
        row2 = dict(row)
        row2["customer_names"] = " | ".join(names)
        row2["top_customers"] = " | ".join(names[:15])
        row2["customer_list_count"] = str(len(names))
        out.append(row2)
    return out


class AppHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, directory=None, **kwargs):
        raw_items = [normalize_item(r) for r in load_csv(DATA / "roadmap_unified.csv")]
        self.roadmap_items = apply_row_edits(raw_items, load_edits())
        self.cluster_rows = load_csv(DATA / "salesforce_semantic_clusters.csv")
        if not self.cluster_rows:
            self.cluster_rows = load_csv(DATA / "salesforce_clusters_seed_ranked.csv")
        self.assignment_rows = load_csv(DATA / "salesforce_semantic_assignments.csv")
        self.request_rows = load_csv(DATA / "salesforce_requests_unified.csv")
        self.cluster_rows = enrich_clusters_with_customers(self.cluster_rows, self.assignment_rows, self.request_rows)
        # Merge in manually added cluster rows + filter deleted IDs
        cedits = load_cluster_edits()
        deleted_ids = set(cedits.get("deleted_ids", []))
        manual_rows = [r for r in cedits.get("manual_rows", {}).values()
                       if r.get("cluster_id") not in deleted_ids]
        if deleted_ids:
            self.cluster_rows = [r for r in self.cluster_rows
                                 if (r.get("cluster_id") or "") not in deleted_ids]
        if manual_rows:
            self.cluster_rows = list(self.cluster_rows) + manual_rows
        self.shipped_rows = load_csv(DATA / "shipped_detection_results.csv")
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def current_user_email(self):
        cookie_header = self.headers.get("Cookie", "")
        cookies = parse_cookies(cookie_header)
        session = cookies.get(SESSION_COOKIE, "")
        body = verify_session(session)
        if not body:
            return ""
        email = normalize_email(body.get("email"))
        ensure_bootstrap_admin(email)
        return email

    def current_user_role(self):
        email = self.current_user_email()
        if not email:
            return ROLE_VIEW
        user = get_user_record(email)
        if not user:
            return ROLE_VIEW
        return normalize_role(user.get("role", ROLE_VIEW))

    def log_activity(self, action, details=None, target_row_id=""):
        email = self.current_user_email() or "anonymous"
        data = load_activity()
        events = data.setdefault("events", [])
        events.append({
            "ts": int(time.time()),
            "actor_email": email,
            "action": action,
            "target_row_id": target_row_id or "",
            "details": details or "",
        })
        save_activity(data)

    def do_GET(self):
        parsed = urlparse(self.path)
        if not self.is_authorized(parsed.path):
            if parsed.path.startswith("/api/"):
                return self.respond_json({"error": "Unauthorized"}, status=HTTPStatus.UNAUTHORIZED)
            self.send_response(HTTPStatus.FOUND)
            self.send_header("Location", "/dashboard/login.html")
            self.end_headers()
            return
        if parsed.path.startswith("/api/"):
            self.handle_api(parsed)
            return
        return super().do_GET()

    def do_OPTIONS(self):
        self.send_response(HTTPStatus.OK)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, x-workspace-id")
        self.send_header("Content-Length", "0")
        self.end_headers()

    def do_POST(self):
        parsed = urlparse(self.path)
        if not parsed.path.startswith("/api/"):
            return self.respond_json({"error": "Not found"}, status=HTTPStatus.NOT_FOUND)
        if not self.is_authorized(parsed.path):
            return self.respond_json({"error": "Unauthorized"}, status=HTTPStatus.UNAUTHORIZED)

        # Handle file upload separately (multipart/form-data)
        if parsed.path == "/api/sf-data/upload":
            return self.handle_sf_data_upload()

        # For all other endpoints, parse JSON payload
        length = int(self.headers.get("Content-Length", "0") or 0)
        raw = self.rfile.read(length) if length > 0 else b"{}"
        try:
            payload = json.loads(raw.decode("utf-8"))
        except Exception:
            return self.respond_json({"error": "Invalid JSON"}, status=HTTPStatus.BAD_REQUEST)

        if parsed.path == "/api/auth/register":
            email = normalize_email(payload.get("email"))
            password = str(payload.get("password") or "")
            invite_token = str(payload.get("invite_token") or "")
            if not email_is_valid(email):
                return self.respond_json({"ok": False, "error": "Valid email is required"}, status=HTTPStatus.BAD_REQUEST)
            if len(password) < 8:
                return self.respond_json({"ok": False, "error": "Password must be at least 8 characters"}, status=HTTPStatus.BAD_REQUEST)

            users = load_users()
            users.setdefault("users", {})
            if email in users["users"]:
                return self.respond_json({"ok": False, "error": "Email already registered"}, status=HTTPStatus.CONFLICT)

            hp = hash_password(password)
            users["users"][email] = {
                "email": email,
                "salt": hp["salt"],
                "password_hash": hp["hash"],
                "role": ROLE_VIEW,
                "created_at": int(time.time()),
            }
            save_users(users)
            ensure_bootstrap_admin(email)
            applied = apply_invite_to_user(email, invite_token)
            if applied:
                self.log_activity("invite_accepted", details=f"{email} accepted invite as {normalize_role(applied.get('role'))}")

            session = issue_session(email)
            self.send_response(HTTPStatus.OK)
            self.send_header("Set-Cookie", f"{SESSION_COOKIE}={session}; Path=/; HttpOnly; SameSite=Lax")
            self.send_header("Content-Type", "application/json; charset=utf-8")
            body = json.dumps({"ok": True, "email": email}, ensure_ascii=True).encode("utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            self.log_activity("register", details=f"New account registered: {email}")
            return

        if parsed.path == "/api/auth/login":
            email = normalize_email(payload.get("email"))
            password = str(payload.get("password") or "")
            invite_token = str(payload.get("invite_token") or "")
            users = load_users().get("users", {})
            user = users.get(email)
            if not user:
                return self.respond_json({"ok": False, "error": "Invalid credentials"}, status=HTTPStatus.UNAUTHORIZED)
            if not verify_password(password, user.get("salt", ""), user.get("password_hash", "")):
                return self.respond_json({"ok": False, "error": "Invalid credentials"}, status=HTTPStatus.UNAUTHORIZED)
            ensure_bootstrap_admin(email)
            applied = apply_invite_to_user(email, invite_token)
            if applied:
                self.log_activity("invite_accepted", details=f"{email} accepted invite as {normalize_role(applied.get('role'))}")

            session = issue_session(email)
            self.send_response(HTTPStatus.OK)
            self.send_header("Set-Cookie", f"{SESSION_COOKIE}={session}; Path=/; HttpOnly; SameSite=Lax")
            self.send_header("Content-Type", "application/json; charset=utf-8")
            body = json.dumps({"ok": True, "email": email}, ensure_ascii=True).encode("utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            self.log_activity("login", details=f"User logged in: {email}")
            return

        if parsed.path == "/api/auth/google":
            id_token = (payload.get("id_token") or "").strip()
            result = verify_google_id_token(id_token)
            if not result.get("ok"):
                return self.respond_json({"ok": False, "error": result.get("reason", "Login failed")}, status=HTTPStatus.UNAUTHORIZED)
            email = result["email"]
            session = issue_session(email)
            secure = "; Secure" if self.headers.get("X-Forwarded-Proto", "http") == "https" else ""
            cookie = f"{SESSION_COOKIE}={session}; Path=/; HttpOnly; SameSite=Lax{secure}"
            self.send_response(HTTPStatus.OK)
            self.send_header("Set-Cookie", cookie)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            body = json.dumps({"ok": True, "email": email}, ensure_ascii=True).encode("utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        if parsed.path == "/api/auth/logout":
            email = self.current_user_email()
            self.send_response(HTTPStatus.OK)
            self.send_header("Set-Cookie", f"{SESSION_COOKIE}=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax")
            self.send_header("Content-Type", "application/json; charset=utf-8")
            body = json.dumps({"ok": True}, ensure_ascii=True).encode("utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            if email:
                self.log_activity("logout", details=f"User logged out: {email}")
            return

        if parsed.path == "/api/invitations/create":
            actor_email = self.current_user_email()
            actor_role = self.current_user_role()
            if not actor_email:
                return self.respond_json({"ok": False, "error": "Unauthorized"}, status=HTTPStatus.UNAUTHORIZED)
            invite_email = normalize_email(payload.get("email"))
            requested_role = normalize_role(payload.get("role", ROLE_VIEW))
            if not email_is_valid(invite_email):
                return self.respond_json({"ok": False, "error": "Valid invite email is required"}, status=HTTPStatus.BAD_REQUEST)
            effective_role = requested_role if role_can_manage_users(actor_role) else ROLE_VIEW
            token = secrets.token_urlsafe(24)
            invites = load_invites()
            invites.setdefault("invites", {})
            invites["invites"][token] = {
                "token": token,
                "email": invite_email,
                "role": effective_role,
                "requested_role": requested_role,
                "invited_by": actor_email,
                "invited_by_role": actor_role,
                "status": "pending",
                "created_at": int(time.time()),
            }
            save_invites(invites)
            invite_link = f"/dashboard/login.html?invite_token={token}&email={invite_email}"
            self.log_activity("invite_user", details=f"Invited {invite_email} as {effective_role}")
            return self.respond_json({"ok": True, "email": invite_email, "role": effective_role, "invite_link": invite_link})

        if parsed.path == "/api/users/set-role":
            actor_role = self.current_user_role()
            if not role_can_manage_users(actor_role):
                return self.respond_json({"ok": False, "error": "Only admin can change roles"}, status=HTTPStatus.FORBIDDEN)
            email = normalize_email(payload.get("email"))
            role = normalize_role(payload.get("role"))
            if not email_is_valid(email):
                return self.respond_json({"ok": False, "error": "Valid email is required"}, status=HTTPStatus.BAD_REQUEST)
            users = load_users()
            users.setdefault("users", {})
            user = users["users"].get(email)
            if not user:
                return self.respond_json({"ok": False, "error": "User not found"}, status=HTTPStatus.NOT_FOUND)
            user["role"] = role
            users["users"][email] = user
            save_users(users)
            self.log_activity("set_role", details=f"Set role {role} for {email}")
            return self.respond_json({"ok": True, "email": email, "role": role})

        if parsed.path == "/api/clusters/add-row":
            if not role_can_edit(self.current_user_role()):
                return self.respond_json({"ok": False, "error": "Only admin/edit users can add cluster items"}, status=HTTPStatus.FORBIDDEN)
            fields = payload.get("fields") or {}
            if not isinstance(fields, dict):
                return self.respond_json({"error": "fields must be an object"}, status=HTTPStatus.BAD_REQUEST)
            cluster_id = f"MANUAL-{int(time.time() * 1000)}"
            safe = {k: ("" if v is None else str(v)) for k, v in fields.items()}
            safe["cluster_id"] = cluster_id
            safe["manual"] = "true"
            # Auto-set rank to end of list
            cedits = load_cluster_edits()
            cedits.setdefault("manual_rows", {})
            cedits["manual_rows"][cluster_id] = safe
            save_cluster_edits(cedits)
            self.log_activity("add_cluster", details=f"Manually added cluster: {safe.get('cluster_label','?')}")
            return self.respond_json({"ok": True, "cluster_id": cluster_id})

        if parsed.path == "/api/clusters/delete-row":
            if not role_can_edit(self.current_user_role()):
                return self.respond_json({"ok": False, "error": "Only admin/edit users can delete cluster items"}, status=HTTPStatus.FORBIDDEN)
            cluster_id = (payload.get("cluster_id") or "").strip()
            if not cluster_id:
                return self.respond_json({"error": "cluster_id is required"}, status=HTTPStatus.BAD_REQUEST)
            cedits = load_cluster_edits()
            cedits.setdefault("manual_rows", {})
            cedits.setdefault("deleted_ids", [])
            # Remove from manual rows if present, otherwise add to deleted_ids blacklist
            if cluster_id in cedits["manual_rows"]:
                del cedits["manual_rows"][cluster_id]
            else:
                if cluster_id not in cedits["deleted_ids"]:
                    cedits["deleted_ids"].append(cluster_id)
            save_cluster_edits(cedits)
            self.log_activity("delete_cluster", details=f"Deleted cluster: {cluster_id}")
            return self.respond_json({"ok": True, "cluster_id": cluster_id})

        if parsed.path == "/api/extract-document":
            # Accepts { filename, mime_type, data_b64 } — returns { text, error }
            filename = str(payload.get("filename") or "")
            mime_type = str(payload.get("mime_type") or "")
            data_b64 = str(payload.get("data_b64") or "")
            if not data_b64:
                return self.respond_json({"ok": False, "error": "data_b64 is required"}, status=HTTPStatus.BAD_REQUEST)
            try:
                # data_b64 may arrive as a full data URL (data:...;base64,XXXX) or raw base64
                if "," in data_b64:
                    data_b64 = data_b64.split(",", 1)[1]
                data_bytes = base64.b64decode(data_b64)
            except Exception as ex:
                return self.respond_json({"ok": False, "error": f"Invalid base64: {ex}"}, status=HTTPStatus.BAD_REQUEST)
            text, err = extract_text_from_document(data_bytes, filename, mime_type)
            return self.respond_json({"ok": True, "text": text, "error": err or None})

        if parsed.path == "/api/roadmap/save-row":
            if not role_can_edit(self.current_user_role()):
                return self.respond_json({"ok": False, "error": "Only admin/edit users can edit roadmap items"}, status=HTTPStatus.FORBIDDEN)
            row_id = (payload.get("row_id") or "").strip()
            fields = payload.get("fields") or {}
            if not row_id or not isinstance(fields, dict):
                return self.respond_json({"error": "row_id and fields are required"}, status=HTTPStatus.BAD_REQUEST)
            edits = load_edits()
            edits.setdefault("rows", {})
            edits["rows"].setdefault(row_id, {})
            for k, v in fields.items():
                edits["rows"][row_id][k] = "" if v is None else str(v)
            save_edits(edits)
            self.log_activity("update_row", details="Updated roadmap item fields", target_row_id=row_id)
            return self.respond_json({"ok": True, "row_id": row_id})

        if parsed.path == "/api/roadmap/save-batch":
            if not role_can_edit(self.current_user_role()):
                return self.respond_json({"ok": False, "error": "Only admin/edit users can reorder roadmap items"}, status=HTTPStatus.FORBIDDEN)
            items = payload.get("items") or []
            if not isinstance(items, list):
                return self.respond_json({"error": "items must be a list"}, status=HTTPStatus.BAD_REQUEST)
            edits = load_edits()
            edits.setdefault("rows", {})
            for entry in items:
                if not isinstance(entry, dict):
                    continue
                row_id = (entry.get("row_id") or "").strip()
                fields = entry.get("fields") or {}
                if not row_id or not isinstance(fields, dict):
                    continue
                edits["rows"].setdefault(row_id, {})
                for k, v in fields.items():
                    edits["rows"][row_id][k] = "" if v is None else str(v)
            save_edits(edits)
            self.log_activity("reorder_rows", details=f"Updated {len(items)} rows")
            return self.respond_json({"ok": True, "count": len(items)})

        if parsed.path == "/api/roadmap/add-row":
            if not role_can_edit(self.current_user_role()):
                return self.respond_json({"ok": False, "error": "Only admin/edit users can add roadmap items"}, status=HTTPStatus.FORBIDDEN)
            fields = payload.get("fields") or {}
            if not isinstance(fields, dict):
                return self.respond_json({"error": "fields must be an object"}, status=HTTPStatus.BAD_REQUEST)

            row_id = f"MANUAL-{int(time.time() * 1000)}"
            edits = load_edits()
            edits.setdefault("added_rows", {})
            safe_fields = {}
            for k, v in fields.items():
                safe_fields[k] = "" if v is None else str(v)
            safe_fields["roadmap_item_id"] = row_id
            edits["added_rows"][row_id] = safe_fields
            save_edits(edits)
            self.log_activity("add_row", details=f"Added roadmap item: {safe_fields.get('title', 'Untitled')}", target_row_id=row_id)
            return self.respond_json({"ok": True, "row_id": row_id})

        if parsed.path == "/api/roadmap/delete-row":
            row_id = (payload.get("row_id") or "").strip()
            if not row_id:
                return self.respond_json({"ok": False, "error": "row_id is required"}, status=HTTPStatus.BAD_REQUEST)
            actor = self.current_user_email()
            if normalize_email(actor) != DELETE_ADMIN_EMAIL:
                return self.respond_json({"ok": False, "error": f"Only {DELETE_ADMIN_EMAIL} can delete roadmap items"}, status=HTTPStatus.FORBIDDEN)

            edits = load_edits()
            edits.setdefault("deleted_rows", {})
            edits["deleted_rows"][row_id] = {"deleted_at": int(time.time()), "deleted_by": actor}
            save_edits(edits)
            self.log_activity("delete_row", details=f"Deleted roadmap item by admin {actor}", target_row_id=row_id)
            return self.respond_json({"ok": True, "row_id": row_id})

        if parsed.path == "/api/sf-proxy/generate":
            return self.proxy_sf_generate(payload)

        if parsed.path == "/api/sf-data/process":
            filename = payload.get("filename", "")
            return self.handle_sf_data_process(filename)

        return self.respond_json({"error": "Not found"}, status=HTTPStatus.NOT_FOUND)

    def handle_api(self, parsed):
        params = parse_qs(parsed.query)
        path = parsed.path

        if path == "/api/auth/config":
            return self.respond_json({
                "google_client_id": GOOGLE_CLIENT_ID,
                "allowed_domain": ALLOWED_DOMAIN,
                "auth_disabled": AUTH_DISABLED,
            })

        if path == "/api/auth/me":
            email = self.current_user_email()
            if not email:
                return self.respond_json({"ok": False, "error": "Unauthorized"}, status=HTTPStatus.UNAUTHORIZED)
            role = self.current_user_role()
            return self.respond_json({
                "ok": True,
                "email": email,
                "role": role,
                "permissions": {
                    "can_add": role_can_edit(role),
                    "can_edit": role_can_edit(role),
                    "can_delete": normalize_email(email) == DELETE_ADMIN_EMAIL,
                    "can_manage_users": role_can_manage_users(role),
                    "can_invite": True,
                },
            })

        if path == "/api/users/list":
            if not role_can_manage_users(self.current_user_role()):
                return self.respond_json({"ok": False, "error": "Only admin can manage users"}, status=HTTPStatus.FORBIDDEN)
            users = load_users().get("users", {})
            items = []
            for email, user in users.items():
                items.append({
                    "email": normalize_email(email),
                    "role": normalize_role(user.get("role", ROLE_VIEW)),
                    "created_at": user.get("created_at", 0),
                })
            items.sort(key=lambda x: x["email"])
            return self.respond_json({"ok": True, "items": items})

        if path == "/api/notifications":
            limit = int(normalize_query_value(params, "limit", default="30"))
            me = self.current_user_email()
            events = load_activity().get("events", [])
            events = list(reversed(events))
            notifications = []
            for e in events:
                actor = normalize_email(e.get("actor_email", ""))
                if me and actor == me:
                    continue
                notifications.append(e)
                if len(notifications) >= limit:
                    break
            return self.respond_json({"items": notifications, "count": len(notifications)})

        if path == "/api/roadmap/items":
            product = normalize_query_value(params, "product")
            quarter = normalize_query_value(params, "quarter")
            status = normalize_query_value(params, "status")
            rows = filter_items(self.roadmap_items, product=product, quarter=quarter, status=status)
            return self.respond_json({"items": rows})

        if path == "/api/roadmap/products":
            return self.respond_json({"products": sorted({x["product"] for x in self.roadmap_items})})

        if path == "/api/roadmap/statuses":
            statuses = sorted({x["status_raw"] for x in self.roadmap_items if x["status_raw"]})
            return self.respond_json({"statuses": statuses})

        if path == "/api/roadmap/summary":
            product = normalize_query_value(params, "product")
            return self.respond_json(roadmap_summary(self.roadmap_items, product=product))

        if path == "/api/roadmap/product-rollups":
            return self.respond_json({"items": product_rollups(self.roadmap_items)})

        if path == "/api/roadmap/quarterly":
            product = normalize_query_value(params, "product")
            return self.respond_json({"items": roadmap_quarterly(self.roadmap_items, product=product)})

        if path == "/api/clusters":
            limit = int(normalize_query_value(params, "limit", default="50"))
            return self.respond_json({"items": self.cluster_rows[:limit]})

        # /api/clusters/<cluster_id>/items  — individual tickets for one cluster
        if path.startswith("/api/clusters/") and path.endswith("/items"):
            cluster_id = path[len("/api/clusters/"):-len("/items")]
            req_ids = [
                a.get("request_id", "").strip()
                for a in (self.assignment_rows or [])
                if (a.get("cluster_id") or "").strip() == cluster_id
            ]
            req_by_id = {}
            for r in (self.request_rows or []):
                rid = (r.get("Request ID") or r.get("request_id") or "").strip()
                if rid:
                    req_by_id[rid] = r
            items = []
            for req_id in req_ids:
                r = req_by_id.get(req_id)
                if r:
                    items.append({
                        "request_id": req_id,
                        "case_number": r.get("Case Number") or r.get("casenumber") or "",
                        "title": r.get("Title") or r.get("title") or r.get("subject") or "",
                        "description": r.get("Description") or r.get("description") or "",
                        "account_name": r.get("Account Name") or r.get("account_name") or "",
                        "status": r.get("Status") or r.get("status") or "",
                        "issue_type_1": r.get("Issue Type 1") or r.get("issue_type_1") or "",
                        "issue_type_2": r.get("Issue Type 2") or r.get("issue_type_2") or "",
                        "severity": r.get("Severity") or r.get("severity") or "",
                        "priority": r.get("Priority") or r.get("priority") or "",
                        "created_date": r.get("Created Date") or r.get("createddate") or "",
                    })
            return self.respond_json({"items": items, "total": len(items)})

        if path == "/api/shipped":
            limit = int(normalize_query_value(params, "limit", default="50"))
            return self.respond_json({"items": self.shipped_rows[:limit]})

        if path.startswith("/api/sf-proxy/download/"):
            file_id = path.replace("/api/sf-proxy/download/", "")
            return self.proxy_sf_download(file_id)

        return self.respond_json({"error": "Not found"}, status=HTTPStatus.NOT_FOUND)

    def handle_sf_data_upload(self):
        """Handle Salesforce data Excel file upload"""
        try:
            # Get content type and boundary
            content_type = self.headers.get("Content-Type", "")
            if not content_type.startswith("multipart/form-data"):
                return self.respond_json({"error": "Content-Type must be multipart/form-data"}, status=HTTPStatus.BAD_REQUEST)
            
            # Extract boundary
            boundary = None
            for part in content_type.split(";"):
                part = part.strip()
                if part.startswith("boundary="):
                    boundary = part.split("=", 1)[1]
                    break
            
            if not boundary:
                return self.respond_json({"error": "No boundary found in Content-Type"}, status=HTTPStatus.BAD_REQUEST)
            
            # Read the entire body
            content_length = int(self.headers.get("Content-Length", 0))
            if content_length == 0:
                return self.respond_json({"error": "No content in request"}, status=HTTPStatus.BAD_REQUEST)
            
            body = self.rfile.read(content_length)
            
            # Parse multipart data manually
            boundary_marker = f"--{boundary}".encode()
            parts = body.split(boundary_marker)
            
            file_data = None
            filename = "salesforce_data.xlsx"
            
            for part in parts:
                if not part or part == b'--\r\n' or part == b'--':
                    continue
                
                # Look for file content
                if b'filename=' in part:
                    # Split headers and content
                    try:
                        header_end = part.index(b'\r\n\r\n')
                        headers_section = part[:header_end].decode('utf-8', errors='ignore')
                        content = part[header_end + 4:]
                        
                        # Extract filename
                        if 'filename="' in headers_section:
                            start = headers_section.index('filename="') + 10
                            end = headers_section.index('"', start)
                            filename = headers_section[start:end]
                        
                        # Remove trailing \r\n if present
                        if content.endswith(b'\r\n'):
                            content = content[:-2]
                        
                        file_data = content
                        break
                    except (ValueError, IndexError):
                        continue
            
            if not file_data:
                return self.respond_json({"error": "No file data found in upload"}, status=HTTPStatus.BAD_REQUEST)
            
            # Save file to data/raw directory
            raw_dir = DATA.parent / "raw"
            raw_dir.mkdir(exist_ok=True)
            
            # Use a timestamped filename to avoid conflicts
            timestamp = int(time.time())
            safe_filename = f"salesforce_upload_{timestamp}.xlsx"
            file_path = raw_dir / safe_filename
            
            with open(file_path, "wb") as f:
                f.write(file_data)
            
            return self.respond_json({
                "ok": True,
                "message": "File uploaded successfully",
                "filename": safe_filename,
                "path": str(file_path),
                "size": len(file_data),
                "original_filename": filename
            })
            
        except Exception as e:
            import traceback
            return self.respond_json({
                "error": f"Upload failed: {str(e)}",
                "traceback": traceback.format_exc()
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)

    def handle_sf_data_process(self, filename):
        """Process uploaded Salesforce data using normalization and LLM scripts"""
        import subprocess
        import shutil
        
        try:
            if not filename:
                return self.respond_json({"error": "Filename is required"}, status=HTTPStatus.BAD_REQUEST)
            
            # Verify file exists
            raw_dir = DATA.parent / "raw"
            file_path = raw_dir / filename
            
            if not file_path.exists():
                return self.respond_json({"error": "Uploaded file not found"}, status=HTTPStatus.NOT_FOUND)
            
            # Copy uploaded file to replace the default Salesforce file
            # The normalize script expects report1772101003441.xls
            target_file = raw_dir / "report1772101003441.xls"
            try:
                shutil.copy2(file_path, target_file)
            except Exception as e:
                return self.respond_json({
                    "error": f"Failed to copy file: {str(e)}"
                }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
            
            # Step 1: Run normalize_inputs.py to process raw data
            normalize_script = ROOT / "scripts" / "normalize_inputs.py"
            
            if not normalize_script.exists():
                return self.respond_json({"error": "Normalization script not found"}, status=HTTPStatus.NOT_FOUND)
            
            normalize_result = subprocess.run(
                ["python3", str(normalize_script)],
                cwd=str(ROOT),
                capture_output=True,
                text=True,
                timeout=120  # 2 minute timeout
            )
            
            if normalize_result.returncode != 0:
                error_msg = normalize_result.stderr or normalize_result.stdout or "Unknown error"
                return self.respond_json({
                    "error": f"Normalization failed: {error_msg}",
                    "returncode": normalize_result.returncode
                }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
            
            # Step 2: Run semantic clustering LLM script
            cluster_script = ROOT / "scripts" / "semantic_cluster_llm.py"

            if not cluster_script.exists():
                return self.respond_json({"error": "Clustering script not found"}, status=HTTPStatus.NOT_FOUND)

            cluster_cmd = ["python3", str(cluster_script)]
            openai_api_key = os.environ.get("OPENAI_API_KEY", "").strip()
            if openai_api_key:
                cluster_cmd += ["--use-llm", "--deep-think"]

            cluster_env = os.environ.copy()
            cluster_timeout = 600 if openai_api_key else 300  # 10 min with LLM, 5 min without

            cluster_result = subprocess.run(
                cluster_cmd,
                cwd=str(ROOT),
                capture_output=True,
                text=True,
                env=cluster_env,
                timeout=cluster_timeout
            )
            
            if cluster_result.returncode != 0:
                error_msg = cluster_result.stderr or cluster_result.stdout or "Unknown error"
                return self.respond_json({
                    "error": f"Clustering failed: {error_msg}",
                    "returncode": cluster_result.returncode,
                    "normalize_output": normalize_result.stdout
                }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
            
            # Step 3: Run shipped detection against ClearTax docs + release notes
            shipped_script = ROOT / "scripts" / "shipped_detection.py"
            shipped_output = ""
            shipped_warning = None
            if shipped_script.exists():
                shipped_cmd = ["python3", str(shipped_script)]
                if openai_api_key:
                    shipped_cmd += ["--use-llm"]
                try:
                    shipped_result = subprocess.run(
                        shipped_cmd,
                        cwd=str(ROOT),
                        capture_output=True,
                        text=True,
                        env=cluster_env,
                        timeout=300
                    )
                    if shipped_result.returncode == 0:
                        shipped_output = shipped_result.stdout
                    else:
                        shipped_warning = f"Shipped detection completed with warnings: {shipped_result.stderr or shipped_result.stdout}"
                except subprocess.TimeoutExpired:
                    shipped_warning = "Shipped detection timed out; results may be incomplete."
                except Exception as e:
                    shipped_warning = f"Shipped detection skipped: {str(e)}"

            response_payload = {
                "ok": True,
                "message": "Salesforce data processed successfully. Feature request raised on salesforce have been updated.",
                "normalize_output": normalize_result.stdout,
                "cluster_output": cluster_result.stdout,
                "shipped_output": shipped_output,
                "files_updated": [
                    "data/processed/salesforce_requests_unified.csv",
                    "data/processed/salesforce_semantic_clusters.csv",
                    "data/processed/salesforce_semantic_assignments.csv",
                    "data/processed/shipped_detection_results.csv",
                ]
            }
            if shipped_warning:
                response_payload["shipped_warning"] = shipped_warning
            return self.respond_json(response_payload)
            
        except subprocess.TimeoutExpired:
            return self.respond_json({"error": "Processing timeout (exceeded allowed time)"}, status=HTTPStatus.REQUEST_TIMEOUT)
        except Exception as e:
            import traceback
            return self.respond_json({
                "error": f"Processing failed: {str(e)}",
                "traceback": traceback.format_exc()
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)

    def proxy_sf_generate(self, payload):
        """Proxy request to Salesforce API to generate Excel file"""
        sf_base_url = os.environ.get("SF_PROXY_BASE_URL", "http://localhost:8080").rstrip("/")
        try:
            outbound_payload = dict(payload or {})
            now_utc = datetime.utcnow()
            from_utc = now_utc - timedelta(days=183)
            # Send explicit 6-month hints for upstream implementations that support different field names.
            outbound_payload["last_n_months"] = 6
            outbound_payload["months"] = 6
            outbound_payload["from_date"] = from_utc.strftime("%Y-%m-%d")
            outbound_payload["to_date"] = now_utc.strftime("%Y-%m-%d")

            sf_url = f"{sf_base_url}/public/metrics/getSalesForceDataAsExcel"
            req_data = json.dumps(outbound_payload).encode("utf-8")
            req = request.Request(
                sf_url,
                data=req_data,
                headers={
                    "x-workspace-id": "3721d2ac-2b59-4421-a6cb-5ad7ac4faa01",
                    "Content-Type": "application/json"
                },
                method="POST"
            )
            with request.urlopen(req, timeout=30) as response:
                response_data = json.loads(response.read().decode("utf-8"))
                # Rewrite the downloadUrl to use our proxy
                if "downloadUrl" in response_data:
                    original_url = response_data["downloadUrl"]
                    file_id = original_url.split("/")[-1]
                    response_data["downloadUrl"] = f"/api/sf-proxy/download/{file_id}"
                return self.respond_json(response_data)
        except Exception as e:
            # Always fall back to local processed Salesforce data.
            filename = f"salesforce_data_{int(time.time())}.xlsx"
            row_count = len(self.request_rows)
            return self.respond_json({
                "ok": True,
                "downloadUrl": "/api/sf-proxy/download/local-salesforce",
                "fileName": filename,
                "fileSizeBytes": 0,
                "expiresIn": "session",
                "rowCount": row_count,
                "warning": f"Upstream Salesforce service unavailable. Serving {row_count} locally processed rows.",
            })

    def proxy_sf_download(self, file_id):
        """Proxy request to download Excel file from Salesforce API"""
        if file_id == "local-salesforce":
            # Build rows from disk CSV (always freshest) or fall back to in-memory.
            sf_csv_path = DATA / "salesforce_requests_unified.csv"
            if sf_csv_path.exists():
                with open(sf_csv_path, newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    fieldnames = reader.fieldnames or []
                    rows = list(reader)
            else:
                rows = self.request_rows or []
                fieldnames = []
                seen = set()
                for row in rows:
                    for key in row.keys():
                        if key not in seen:
                            seen.add(key)
                            fieldnames.append(key)
                if not fieldnames:
                    fieldnames = [
                        "Request ID", "Case Number", "Created Date", "Status", "Title",
                        "Description", "Account Name", "Product", "Priority",
                        "Ticket Count", "Account Active ARR",
                    ]

            # Convert to xlsx using openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Salesforce Data"

            # Header row (bold)
            from openpyxl.styles import Font
            header_font = Font(bold=True)
            ws.append(list(fieldnames))
            for cell in ws[1]:
                cell.font = header_font

            # Data rows
            for row in rows:
                ws.append([row.get(col, "") for col in fieldnames])

            # Auto-fit column widths (approximate)
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

            buf = io.BytesIO()
            wb.save(buf)
            content = buf.getvalue()

            stamp = int(time.time())
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="salesforce_data_{stamp}.xlsx"')
            self.send_header("Content-Length", str(len(content)))
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(content)
            return

        try:
            sf_base_url = os.environ.get("SF_PROXY_BASE_URL", "http://localhost:8080").rstrip("/")
            sf_url = f"{sf_base_url}/public/metrics/downloadExcel/{file_id}"
            req = request.Request(sf_url, method="GET")
            with request.urlopen(req, timeout=30) as response:
                content = response.read()
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", 'attachment; filename="salesforce_data.xlsx"')
                self.send_header("Content-Length", str(len(content)))
                self.send_header("Access-Control-Allow-Origin", "*")
                self.end_headers()
                self.wfile.write(content)
        except Exception as e:
            return self.respond_json({"error": f"Failed to download file: {str(e)}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)

    def respond_json(self, payload, status=HTTPStatus.OK):
        body = json.dumps(payload, ensure_ascii=True).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, x-workspace-id")
        self.end_headers()
        self.wfile.write(body)

    def is_authorized(self, path):
        if AUTH_DISABLED:
            return True
        public_paths = {
            "/dashboard/login.html",
            "/dashboard/login.js",
            "/api/auth/config",
            "/api/auth/me",
            "/api/auth/register",
            "/api/auth/login",
            "/api/auth/google",
            "/api/auth/logout",
        }
        if path in public_paths:
            return True
        if path.startswith("/dashboard/login"):
            return True
        if path == "/":
            return False
        if path.startswith("/dashboard/") or path.startswith("/api/"):
            cookie_header = self.headers.get("Cookie", "")
            cookies = parse_cookies(cookie_header)
            session = cookies.get(SESSION_COOKIE, "")
            return verify_session(session) is not None
        return True


def run_server(host="127.0.0.1", port=8000):
    server = ThreadingHTTPServer((host, port), AppHandler)
    print(f"Serving dashboard + API at http://{host}:{port}")
    print("Endpoints: /api/roadmap/items, /api/roadmap/summary, /api/clusters, /api/shipped")
    server.serve_forever()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run roadmap dashboard API server")
    parser.add_argument("--host", default=os.environ.get("HOST", "127.0.0.1"))
    parser.add_argument("--port", type=int, default=int(os.environ.get("PORT", "8000")))
    args = parser.parse_args()
    run_server(args.host, args.port)
